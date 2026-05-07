from dotenv import load_dotenv
import os
from smtplib import SMTP
from email.mime.multipart import MIMEMultipart
from email import encoders
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
# from email.mime.image import MIMEImage
from openpyxl import load_workbook


load_dotenv()

# 메일 보내기 함수
def send_mail():
    # smtp 접속정보 load
    user_email = os.getenv("EMAIL_USER")
    user_password = os.getenv("EMAIL_PASS")
    smtp_server = os.getenv("SMTP_SVR")
    smtp_port = int(os.getenv("SMTP_PORT"))

    # 엑셀 파일첨부
    msg = MIMEMultipart()
    file_path = "project_2j/test.xlsx"
    if not os.path.exists(file_path):
        print("메일에 첨부하려는 엑셀 파일이 존재하지 않습니다.")
        return
    with open(file_path, "rb") as f:
        part = MIMEBase("application", "octet-stream")
        part.set_payload(f.read())
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", f"attachment; filename={os.path.basename(file_path)}")
    msg.attach(part)

    # 메일내용 - 엑셀에서 기사제목 > 텍스트
    msg["Subject"] = "[환율 정보 전달] 금일 환율 이슈"
    msg["From"] = user_email # 보내는 사람
    msg["To"] = user_email # 받는 사람
    # 엑셀 읽기
    wb = load_workbook(file_path)
    ws = wb.active

    headers = []
    for cell in ws[1]:
        headers.append(cell.value)

    # n_title 열 위치 찾기
    title_index = headers.index("n_title")

    body = """
안녕하세요.
금일 환율 변동사항과 관련 뉴스를 전달드립니다.

[환율 관련 뉴스]

"""
    for row in ws.iter_rows(min_row=2, values_only=True):
        title = row[title_index]
        if title:
            body += f"---{title}---\n"
    msg.attach(MIMEText(body, 'plain'))


    
    # 서버 접속/발송
    try:
        with SMTP(smtp_server, smtp_port) as server:
            server.starttls()
            server.login(user_email, user_password)
            server.send_message(msg)
            print("메일 발송 성공")
            # 발송로그저장? 
            
    except Exception as e:
        print(f"오류 발생 : {e}")





# 함수 실행
if __name__ =="__main__":
    send_mail()

    

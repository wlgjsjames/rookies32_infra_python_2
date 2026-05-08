import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference
from datetime import datetime
import os

class ExcelReportManager:
    def __init__(self):
        # [디자인 가이드]
        self.color_header = PatternFill(start_color="F2F4F7", end_color="F2F4F7", fill_type="solid")
        self.color_drop_bg = PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid") # 연한 녹색 (매수적기)
        self.color_rise_bg = PatternFill(start_color="FCE8E6", end_color="FCE8E6", fill_type="solid") # 연한 빨간색 (보류)
        
        self.font_main = Font(name='Malgun Gothic', size=10, color="202124")
        self.font_header = Font(name='Malgun Gothic', size=10, color="202124", bold=True)
        self.font_link = Font(name='Malgun Gothic', size=10, color="0563C1", underline="single")
        
        self.border = Border(
            left=Side(style="thin", color="DADCE0"),
            right=Side(style="thin", color="DADCE0"),
            top=Side(style="thin", color="DADCE0"),
            bottom=Side(style="thin", color="DADCE0")
        )

    def create_report(self, analysis_details, news_info, history_data=None):
        """
        1. analysis_details: 3번(analyzer)의 결과 
        2. news_info: 1번(scraper)의 뉴스 정보
        3. history_data: 2번(DB)에서 가져온 최근 환율 리스트
        """
        # --- [파일명 업데이트] 실행 시점의 날짜를 포함한 고유 파일명 생성 ---
        today_date = datetime.now().strftime("%Y%m%d")
        file_name = f"환율리포트_{today_date}.xlsx"
        
        # 1. [직구 시뮬레이션] 데이터 가공
        rate_data = []
        for curr, data in analysis_details.items():
            today_rate = data['today']
            rate_data.append({
                '통화명': curr,
                '현재 환율': today_rate,
                '전일 환율': data['yesterday'],
                '변동 수치': data['diff'],
                '변동률(%)': data['percent'],
                '등락 여부': '하락' if data['is_dropped'] else ('상승' if data['diff'] > 0 else '보합'),
                'iPhone 15 Pro ($999)': f"{int(today_rate * 999):,}원",
                'AirPods Pro ($249)': f"{int(today_rate * 249):,}원"
            })

        # 2. [관련 뉴스 기사] 데이터 가공 (mailer.py 연동용)
        news_data = [{
            'n_title': news_info.get('title', '제목 없음'),
            'news_link': news_info.get('link', '#'),
            'created_at': news_info.get('date', datetime.now().strftime("%Y.%m.%d"))
        }]

        with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
            # 시트 배치 (mailer.py가 읽기 쉽도록 뉴스를 첫 번째로)
            pd.DataFrame(news_data).to_excel(writer, index=False, sheet_name='관련 뉴스 기사')
            pd.DataFrame(rate_data).to_excel(writer, index=False, sheet_name='오늘의 환율 현황')
            
            # 히스토리 데이터가 있다면 차트용 데이터 시트 추가
            if history_data:
                pd.DataFrame(history_data).to_excel(writer, index=False, sheet_name='데이터 트렌드')

            # 스타일 적용 및 차트 생성
            self._apply_news_style(writer.sheets['관련 뉴스 기사'])
            self._apply_rate_style(writer.sheets['오늘의 환율 현황'])
            
            if history_data:
                self._add_trend_chart(writer.sheets['오늘의 환율 현황'], writer.sheets['데이터 트렌드'], len(history_data))

        return file_name # 생성된 파일명을 반환하여 6번 main.py가 mailer에 전달할 수 있게 함

    def _apply_news_style(self, sheet):
        """뉴스 시트 디자인"""
        for col in range(1, sheet.max_column + 1):
            column_letter = get_column_letter(col)
            sheet.column_dimensions[column_letter].width = 60 if col == 1 else 30
            
            cell = sheet.cell(row=1, column=col)
            cell.fill = self.color_header
            cell.font = self.font_header
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border

        for row in range(2, sheet.max_row + 1):
            for col in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=row, column=col)
                cell.font = self.font_main if col != 2 else self.font_link
                cell.border = self.border
                cell.alignment = Alignment(horizontal='left', indent=1)

    def _apply_rate_style(self, sheet):
        """환율 시트 [신호등] 디자인"""
        for col in range(1, sheet.max_column + 1):
            column_letter = get_column_letter(col)
            sheet.column_dimensions[column_letter].width = 18
            
            cell = sheet.cell(row=1, column=col)
            cell.fill = self.color_header
            cell.font = self.font_header
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border

        for row in range(2, sheet.max_row + 1):
            status_val = sheet.cell(row=row, column=6).value 
            
            for col in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=row, column=col)
                cell.font = self.font_main
                cell.border = self.border
                cell.alignment = Alignment(horizontal='center')

                # [신호등] 하락(녹색), 상승(빨간색) 배경색 지정
                if status_val == '하락':
                    cell.fill = self.color_drop_bg
                elif status_val == '상승':
                    cell.fill = self.color_rise_bg

    def _add_trend_chart(self, target_sheet, data_sheet, data_count):
        """[데이터 트렌드] 꺾은선 차트 삽입"""
        chart = LineChart()
        chart.title = "최근 7일 환율 추이"
        chart.style = 13
        chart.y_axis.title = "환율 (KRW)"
        chart.x_axis.title = "수집일"
        
        data_ref = Reference(data_sheet, min_col=2, min_row=1, max_row=data_count + 1)
        cats_ref = Reference(data_sheet, min_col=1, min_row=2, max_row=data_count + 1)
        
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        target_sheet.add_chart(chart, "J2")

# --- 테스트 실행 ---
if __name__ == "__main__":
    # 데이터 샘플
    mock_analysis = {"USD": {"today": 1340.5, "yesterday": 1350.0, "diff": -9.5, "percent": "-0.7%", "is_dropped": True}}
    mock_news = {"title": "환율 하락세 지속", "link": "https://kita.net/news/123", "date": "2026.05.08"}
    mock_history = [{"date": "05.07", "USD": 1350}, {"date": "05.08", "USD": 1340}]

    manager = ExcelReportManager()
    output_path = manager.create_report(mock_analysis, mock_news, mock_history)
    print(f"파일 저장 완료: {output_path}")